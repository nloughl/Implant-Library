"""
Knee Implant Reference Library Pipeline - v3
=============================================

Orchestrates the full pipeline:
  1. Bulk download FDA GUDID data          (optimized_gudid_downloader.py)
  2. Bulk download Health Canada MDALL     (mdall_bulk_downloader.py)
  3. Merge both sources on catalogue #     (merge_gudid_mdall.py)
  4. Filter to CJRR manufacturers          (filter_cjrr.py)
  5. Download missing eIFU PDFs            (eifu_downloader.py)   [optional]
  6. Supplement with NZ medical devices CSV (CSVSupplementer)
  7. Enrich materials from eIFU PDFs        (eifu_material_extractor.py)
  8. Classify component types              (ComponentTypeClassifier)
  9. Extract design fields                 (FieldExtractor)
 10. Apply field applicability matrix      (FieldApplicabilityMatrix)
 11. Export enriched ImplantRecord CSV

eIFU file naming convention (used by eifu_downloader.py):
  {manufacturer}_{system_name}_{doc_id}.pdf
  e.g.  stryker_triathlon_QIN4396EIFU.pdf
        zimmer_persona_eIFU.pdf
        depuy_attune_eIFU.pdf

  The manufacturer prefix routes to the correct parser in
  eifu_material_extractor.py.  One PDF covers many catalogue numbers
  (e.g. all femoral sizes of Triathlon) — the extractor handles this.

Usage (full pipeline):
  python implant_lib_v3.py

Usage (skip data downloads, use existing files):
  python implant_lib_v3.py --cjrr cjrr_filtered_20260311_203320.csv

Usage (also download missing eIFUs automatically):
  python implant_lib_v3.py --cjrr cjrr_filtered.csv --download-eifus

Usage (start from a pre-merged file):
  python implant_lib_v3.py --merged merged_gudid_mdall.csv
"""

import argparse
import json
import logging
import re
from dataclasses import dataclass, field, asdict
from datetime import datetime
from enum import Enum
from openpyxl import load_workbook
from openpyxl.styles.numbers import FORMAT_TEXT
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import pandas as pd

# ---------------------------------------------------------------------------
# Pipeline module imports
# ---------------------------------------------------------------------------
from optimized_gudid_downloader import ProductCodeBulkDownloader
from mdall_bulk_downloader import MDALLBulkDownloader
import merge_gudid_mdall as _merger
import filter_cjrr as _cjrr_filter
from eifu_material_extractor import (
    extract_materials as _extract_eifu_materials,
    _detect_variant as _detect_eifu_variant,
)
from eifu_downloader import (
    EIFUDownloadOrchestrator,
    _brand_to_system,
)
from gudid_description_enricher import enrich_dataframe as _enrich_gudid_descriptions
from catalogue_num_sizer import extract_from_catalogue as _cat_size

# ---------------------------------------------------------------------------
# Logging
# ---------------------------------------------------------------------------
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s",
)
logger = logging.getLogger(__name__)


# ===========================================================================
# Deduplication helpers (module-level so static methods can reference them)
# ===========================================================================

# Fields where we coalesce: take the first non-empty value across all rows
# that share a catalogue_num (secondary rows may fill gaps in the primary).
_COALESCE_FIELDS = [
    "component_type",
    "component_description",
    "stability",
    "bearing_type",
    "fixation",
    "metal_material",
    "poly_material",
    "antioxidant",
    "ceramic_material",
    "surface_treatment",
    "side",
    "size",
    "thickness",
    "ap_length",
    "ml_width",
    "diameter",
    "design_articular_surface",
    "design_fixation_surface",
    "device_description",
    "device_sizes_text",
    "brand_name",
    "mdall_brand_name",
    "gmdn_name",
    "notes",
]


def _is_empty(v) -> bool:
    if v is None:
        return True
    if isinstance(v, float) and pd.isna(v):
        return True
    return str(v).strip() == ""


def _merge_catalogue_group(group: pd.DataFrame) -> pd.Series:
    """
    Merge a group of rows that share the same catalogue_num.

    Primary row (iloc[0], already sorted best-first) is the base.
    Empty fields in the primary are filled from secondary rows so that
    information unique to a secondary product code (e.g. bearing_type from
    an NJL mobile-bearing code) is not discarded.
    product_code and device_name are aggregated into pipe-separated strings
    so every associated FDA code is visible in the output.
    """
    primary = group.iloc[0].copy()

    for col in _COALESCE_FIELDS:
        if col not in group.columns:
            continue
        if _is_empty(primary[col]):
            for val in group[col].iloc[1:]:
                if not _is_empty(val):
                    primary[col] = val
                    break

    codes = [str(v) for v in group["product_code"] if not _is_empty(v)]
    primary["product_code"] = " | ".join(dict.fromkeys(codes))

    names = [str(v) for v in group["device_name"] if not _is_empty(v)]
    primary["device_name"] = " | ".join(dict.fromkeys(names))

    return primary


# ===========================================================================
# Data structures
# ===========================================================================


class FieldStatus(Enum):
    PRESENT = "present"
    MISSING = "missing"
    NOT_APPLICABLE = "not_applicable"


@dataclass
class ImplantRecord:
    """Core implant data structure."""

    # --- Identifiers --------------------------------------------------------
    device_id: Optional[str] = None
    catalogue_num: Optional[str] = None
    gtin: Optional[str] = None
    mdall_licence_number: Optional[str] = None
    mdall_device_id: Optional[str] = None

    # --- Basic info ---------------------------------------------------------
    brand_name: Optional[str] = None
    mdall_brand_name: Optional[str] = (
        None  # raw MDALL device name (often richer than GUDID)
    )
    version_model_number: Optional[str] = None
    company_name: Optional[str] = None
    manufacturer: Optional[str] = None  # CJRR canonical name
    product_code: Optional[str] = None
    device_name: Optional[str] = (
        None  # product code description (e.g. "Knee Joint Femoral Component")
    )
    device_description: Optional[str] = (
        None  # GUDID per-device description (e.g. "Triathlon CR Femoral, Size 3")
    )
    device_sizes_text: Optional[str] = None  # GUDID size annotations
    gmdn_name: Optional[str] = None

    # --- Component classification -------------------------------------------
    component_type: Optional[str] = None
    component_description: Optional[str] = None
    replacement_type: str = "P"  # P = primary, R = revision

    # --- Design characteristics ---------------------------------------------
    fixation: Optional[str] = None
    stability: Optional[str] = None
    bearing_type: Optional[str] = (
        None  # Fixed Bearing | Mobile Bearing | Rotating Platform
    )
    side: Optional[str] = None

    # --- Materials (from eIFU / text extraction) ----------------------------
    metal_material: Optional[str] = None
    poly_material: Optional[str] = None
    ceramic_material: Optional[str] = None
    antioxidant: Optional[str] = None  # e.g. "Vitamin E" (Vivacit-E, etc.)
    material_standard: Optional[str] = None

    # --- Dimensions ---------------------------------------------------------
    size: Optional[str] = None
    ap_length: Optional[float] = None
    ml_width: Optional[float] = None
    thickness: Optional[float] = None
    diameter: Optional[float] = None

    # --- Design features ----------------------------------------------------
    design_articular_surface: Optional[str] = None
    design_fixation_surface: Optional[str] = None
    surface_treatment: Optional[str] = (
        None  # fixation-surface coating: HA, PMMA, TiN, TPS, CaP …
    )

    # --- Source tracing -----------------------------------------------------
    data_source: List[str] = field(default_factory=list)
    match_type: Optional[str] = None  # exact | normalized | unmatched
    source_db: Optional[str] = None  # BOTH | GUDID_ONLY | MDALL_ONLY

    # --- Metadata -----------------------------------------------------------
    last_updated: str = field(default_factory=lambda: datetime.now().isoformat())
    confidence_score: float = 0.0
    needs_review: bool = False
    notes: str = ""
    field_status: Dict[str, str] = field(default_factory=dict)


# ===========================================================================
# Utility / enrichment classes
# ===========================================================================


class ComponentTypeClassifier:
    """Classifies implant component type from GMDN terms and product name."""

    @staticmethod
    def _classify_text(text: str) -> Optional[str]:
        t = text.lower()

        if "hip" in t:
            return "Hip Prosthesis"

        if re.search(r"\bstems?\b", t) or "extension stem" in t:
            if "femoral" in t or "femur" in t:
                return "Femoral Stem"
            if "tibial" in t or "tibia" in t:
                return "Tibial Stem"
            return "Stem"

        if any(w in t for w in ["insert", "bearing", "articular surface", "liner"]):
            return "Insert"

        if any(w in t for w in ["patellar", "patella", "patello"]):
            return "Patellar"

        if any(w in t for w in ["tibial", "tibia"]) and "insert" not in t:
            if any(w in t for w in ["baseplate", "tray", "platform", "component"]):
                return "Tibial"

        if any(w in t for w in ["femoral", "femur"]) and not re.search(
            r"\bstems?\b", t
        ):
            if any(w in t for w in ["component", "condyle"]):
                return "Femoral"

        if "femorotibial" in t or ("femoral" in t and "tibial" in t):
            return "TKA System"

        return None

    @staticmethod
    def classify(gmdn_name: str = "", device_description: str = "") -> Optional[str]:
        """Classify component type, checking device_description before gmdn_name.

        device_description is per-device free text (e.g. "Articulating surface,
        vitamin E highly crosslinked UHMWPE, fixed bearing UC") and is more
        specific than the GMDN category label.  gmdn_name is used as fallback
        when device_description is absent or unclassifiable.
        """
        if device_description:
            result = ComponentTypeClassifier._classify_text(device_description)
            if result is not None:
                return result
        return ComponentTypeClassifier._classify_text(gmdn_name)


class FieldExtractor:
    """Extracts structured design fields from free text."""

    @staticmethod
    def extract_fixation(text: str) -> Optional[str]:
        if re.search(r"\bcement(?:ed)?\b|\bnonporous\b", text, re.I):
            return "Cemented"
        if re.search(r"\bcementless\b|\buncemented\b|\bporous\b", text, re.I):
            return "Cementless"
        if re.search(r"\bhybrid\b", text, re.I):
            return "Hybrid"
        return None

    @staticmethod
    def extract_stability(text: str) -> Optional[str]:
        if re.search(r"\bsemi[- ]constrain(?:ed)?\b|\bSC\b", text, re.I):
            return "Semi-Constrained"
        if re.search(
            r"\bPS\b|\bposterior[- ]stabili[sz]ed\b|\bpost stab\b|\bPSR\b|\bP/S\b|\bposterior[- ]sabili[sz]ed\b",
            text,
            re.I,
        ):
            return "PS"
        if re.search(r"\bAS\b|\banterior[- ]stabili[sz]ed?\b", text, re.I):
            return "AS"
        if re.search(r"\bCR\b|\bcruciate[- ]retain(?:ing)?\b", text, re.I):
            return "CR"
        if re.search(
            r"\bmedial\s+pivot\b|\bGMK Sphere\b|\bmedial dished\b|\bMP\b", text, re.I
        ):
            return "Medial Pivot"
        if re.search(r"\bCPS\b", text, re.I):
            return "CPS"
        if re.search(r"\bBCS\b|\bbi[- ]cruciate\b", text, re.I):
            return "BCS"
        if re.search(r"\bCS LIPPED\b|\bCS\b", text, re.I):
            return "CS"
        if re.search(r"\bCCK\b|\bLCCK\b|\bconstrained condylar\b", text, re.I):
            return "CCK"
        if re.search(r"\bTS\b|\btotal stabilizer\b", text, re.I):
            return "TS"
        if re.search(r"\bUC\b|\bdeep[- ]dish\b", text, re.I):
            return "UC"
        if re.search(r"\brotating hinge\b", text, re.I):
            return "Rotating Hinge"
        if re.search(r"\bhinge\b|\bhinged\b", text, re.I):
            return "Hinge"
        if re.search(r"\bconstrain(?:ed)?\b", text, re.I):
            return "Constrained"
        return None

    @staticmethod
    def extract_bearing_type(text: str) -> Optional[str]:
        if re.search(r"\brotating\s+platform\b", text, re.I):
            return "Rotating Platform"
        if re.search(r"\bmobile\s+bearing\b|\bmobile\b|\brotating\b", text, re.I):
            return "Mobile Bearing"
        if re.search(r"\bfixed\s+bearing\b|\bfixed\b", text, re.I):
            return "Fixed Bearing"
        return None

    @staticmethod
    def extract_side(text: str) -> Optional[str]:
        # Medacta tibial hash notation: "#6/12 mm R" — R/L follows the mm value
        m = re.search(r"#\d+/\d+\s*mm\s+([LR])\b", text, re.I)
        if m:
            return "Right" if m.group(1).upper() == "R" else "Left"
        # Medacta femoral suffix: "S1R" / "S2L" — size digit + side letter
        m = re.search(r"\bS\d+([LR])\b", text)
        if m:
            return "Right" if m.group(1).upper() == "R" else "Left"
        if re.search(r"\bleft\b", text, re.I):
            return "Left"
        if re.search(r"\bright\b", text, re.I):
            return "Right"
        if re.search(r"\buniversal\b|\bneutral\b", text, re.I):
            return "Universal"
        return None

    @staticmethod
    def extract_material(
        text: str,
    ) -> Tuple[Optional[str], Optional[str], Optional[str]]:
        """Return (metal, polymer, antioxidant).

        Polymer priority: cross-linking is detected first; Vitamin E is then
        checked independently so that XLPE+VitE products (e.g. Vivacit-E) are
        classified correctly rather than defaulting to UHMWPE+VitE.
        """
        t = text.lower()

        metal = None
        if "cobalt" in t or "cocr" in t or "co-cr" in t:
            metal = "CoCr"
        elif "titanium" in t or "ti-6al-4v" in t:
            metal = "Titanium"
        elif "oxinium" in t or "oxidized zirconium" in t:
            metal = "OxZr"

        polymer = None
        antioxidant = None
        # E-Cross: Medacta brand name for Vitamin E highly cross-linked PE
        if re.search(r"\be[-.]?cross\b", t):
            polymer = "HXLPE+VitE"
            antioxidant = "Vitamin E"
        elif (
            "uhmwpe" in t
            or "polyethylene" in t
            or "standard pe" in t
            or "xlpe" in t
            or bool(re.search(r"\bx3\b", t))
        ):
            is_hxlpe = bool(re.search(r"highly\s+cross.?link(?:ed)?|hxlpe|x3", t))
            is_xlpe = is_hxlpe or bool(re.search(r"xlpe|cross.?link(?:ed)?", t))
            is_vite = bool(re.search(r"vitamin\s*e\b|vit\.?\s*e\b|vivacit", t))

            if is_hxlpe and is_vite:
                polymer = "HXLPE+VitE"
            elif is_hxlpe:
                polymer = "HXLPE"
            elif is_xlpe and is_vite:
                polymer = "XLPE+VitE"
            elif is_xlpe:
                polymer = "XLPE"
            elif is_vite:
                polymer = "UHMWPE+VitE"
            else:
                polymer = "UHMWPE"

            if is_vite:
                antioxidant = "Vitamin E"

        return metal, polymer, antioxidant

    @staticmethod
    def extract_surface_treatment(text: str) -> Optional[str]:
        """Return the fixation-surface coating code, or None.

        PA (Peri-Apatite, Stryker brand name) is normalised to HA.
        Values are checked in specificity order so that, e.g., an HA-coated
        titanium spray surface is reported as HA rather than TPS.
        """
        t = text.lower()
        # Hydroxyapatite — including Peri-Apatite (Stryker) and plain "HA"
        if re.search(
            r"hydroxyapatite|peri.?apatite|\bha\s+coat|\bha\b.*coat|coat.*\bha\b|\+\s*ha\b",
            t,
        ):
            return "HA"
        # Titanium niobium nitride (must precede TiN check — more specific)
        if re.search(r"\btinbn\b|tinbn|titanium\s+niobium\s+nitride", t):
            return "TiNbN"
        # Titanium nitride
        if re.search(r"\btin\b|titanium\s+nitride", t):
            return "TiN"
        # Titanium plasma spray
        if re.search(r"\btps\b|titanium\s+plasma\s+spray|plasma\s+spray", t):
            return "TPS"
        # Calcium phosphate / tricalcium phosphate
        if re.search(r"calcium\s+phosphate|\bca\s*p\b|tricalcium", t):
            return "CaP"
        # PMMA / bone cement (cemented implants with pre-coat)
        if re.search(r"\bpmma\b|polymethylmethacrylate", t):
            return "PMMA"
        return None

    # Written-size normalisation — shared by extract_size patterns 2 & 3.
    # Longer forms must appear before their abbreviations in _W so that the
    # regex alternation doesn't short-circuit (e.g. "SMALL" before "SM").
    _W = r"(?:EXTRA\s+)?(?:SMALL|SML|SM|MEDIUM|MED|STANDARD|STD|LARGE|LGE|LG|XL)\+?(?:\s*/\s*(?:EXTRA\s+)?(?:SMALL|SML|SM|MEDIUM|MED|STANDARD|STD|LARGE|LGE|LG|XL)\+?)?"
    _WRITTEN_NORM: Dict[str, str] = {
        "extra small": "XS",
        "small": "Small",
        "sml": "Small",
        "sm": "Small",
        "medium": "Medium",
        "med": "Medium",
        "standard+": "Standard+",
        "std+": "Standard+",
        "standard": "Standard",
        "std": "Standard",
        "extra large": "XL",
        "large+": "Large+",
        "lge+": "Large+",
        "lg+": "Large+",
        "large": "Large",
        "lge": "Large",
        "lg": "Large",
        "xl": "XL",
    }

    @staticmethod
    def _norm_written_size(s: str) -> str:
        return FieldExtractor._WRITTEN_NORM.get(s.lower().strip(), s.title())

    @staticmethod
    def extract_size(text: str) -> Optional[str]:
        # Priority 0a: hash notation — "#6" from "#6/12 mm" (Medacta) or "#3 Th10" (Corin)
        m = re.search(r"#(\d+)", text)
        if m:
            return m.group(1)
        # Priority 0b: Medacta femoral suffix "S1R" / "S2L" — size digit before side letter
        m = re.search(r"\bS(\d+)[LR]\b", text)
        if m:
            return m.group(1)
        # Priority 1: explicit "Size:" label — numeric ("Size: 13"), letter
        # ("Size: D"), or alphanumeric ("Size: LGE-1"). Space excluded as
        # separator to prevent "SIZE 7-8 10MM" from absorbing the thickness.
        # Negative lookahead skips "Device Size Text, specify:" segments.
        m = re.search(
            r"(?:size|sz)\s*:?\s*(?!text\b|specify\b)([A-Z0-9]+(?:[-/][A-Z0-9]+)*)",
            text,
            re.I,
        )
        if m:
            return m.group(1)
        # Priority 2: parenthesised written size — "(SMALL)", "(LG+)", "(MEDIUM)"
        m = re.search(rf"\(\s*({FieldExtractor._W})\s*\)", text, re.I)
        if m:
            return FieldExtractor._norm_written_size(m.group(1))
        # Priority 3: written size immediately before a mm measurement
        # e.g. "MENISCAL BEARING LG+ 12.5mm" — the written token is the size,
        # the number+mm is the thickness (handled separately by extract_thickness).
        m = re.search(
            rf"\b({FieldExtractor._W})\s+\d+(?:\.\d+)?\s*(?:mm\b|millimeters?\b)",
            text,
            re.I,
        )
        if m:
            return FieldExtractor._norm_written_size(m.group(1))
        return None

    @staticmethod
    def extract_thickness(text: str) -> Optional[str]:
        """Extract insert/component thickness from free text.

        Checks in order:
          0. Medacta "#6/12 mm" — second number after slash is thickness
          0b. Corin "Th10" / "Th 10" label
          1. Labelled form: "Thickness: 10 mm" or "Height: 10 Millimeter"
          2. Numeric size label then bare mm: "SIZE 7-8 10MM" or "SIZE: 5-6 18MM"
          3. Written size token then bare mm: "LG+ 12.5mm"
        Returns the numeric part as a string (e.g. "10", "9.5").
        Plain digit+mm is intentionally avoided to prevent AP/ML widths
        (e.g. "Width: 60 mm") from being misidentified as thickness.
        """
        # Priority 0: Medacta "#6/12 mm" — second number is thickness
        m = re.search(r"#\d+/(\d+(?:\.\d+)?)\s*(?:mm\b|millimeters?\b)", text, re.I)
        if m:
            return m.group(1)
        # Priority 0b: Corin "Th10" or "Th 10" — Th prefix is thickness label
        m = re.search(r"\bTh\s*(\d+(?:\.\d+)?)\b", text)
        if m:
            return m.group(1)
        t = text.lower()
        m = re.search(
            r"(?:thickness|height)\s*:?\s*(\d+(?:\.\d+)?)\s*(?:mm\b|millimeters?\b)", t
        )
        if m:
            return m.group(1)
        m = re.search(
            r"size\s*:?\s*\d+(?:[/-]\d+)?\s+(\d+(?:\.\d+)?)\s*(?:mm\b|millimeters?\b)",
            t,
        )
        if m:
            return m.group(1)
        # Written size token immediately before a bare mm value
        # e.g. "MENISCAL BEARING LG+ 12.5mm" — the number+mm is the thickness.
        _w = r"(?:extra\s+)?(?:small|sml|sm|medium|med|standard|std|large|lge|lg|xl)\+?"
        m = re.search(rf"\b{_w}\s+(\d+(?:\.\d+)?)\s*(?:mm\b|millimeters?\b)", t, re.I)
        if m:
            return m.group(1)
        return None

    @staticmethod
    def extract_dimensions_from_sizes_text(
        sizes_text: str,
    ) -> Dict[str, Optional[float]]:
        """Parse GUDID device_sizes_text pipe-separated key-value pairs.

        Input format: "Width: 60 mm | Height: 9 mm | Length: 72 mm"
        Returns a dict with keys: ap_length, ml_width, thickness, diameter
        (float mm values, or None if the label was not found).
        """
        result: Dict[str, Optional[float]] = {
            "ap_length": None,
            "ml_width": None,
            "thickness": None,
            "diameter": None,
        }
        if not isinstance(sizes_text, str):
            return result
        _MAP = {
            "ap_length": re.compile(r"\b(?:ap|length|anterior.posterior)\b", re.I),
            "ml_width": re.compile(r"\b(?:ml|width|medial.lateral)\b", re.I),
            "thickness": re.compile(r"\b(?:thickness|height)\b", re.I),
            "diameter": re.compile(r"\b(?:diameter|diam)\b", re.I),
        }
        for segment in sizes_text.split("|"):
            segment = segment.strip()
            m_val = re.search(
                r"(\d+(?:\.\d+)?)\s*(?:mm\b|millimeters?\b)", segment, re.I
            )
            if not m_val:
                continue
            val = float(m_val.group(1))
            label = segment[: m_val.start()]
            for field, pat in _MAP.items():
                if pat.search(label) and result[field] is None:
                    result[field] = val
                    break
        return result

    @staticmethod
    def extract_articular_surface(
        text: str, side: Optional[str] = None
    ) -> Optional[str]:
        """Return "Asymmetric" or "Symmetric" from free text, or inferred from side.

        Asymmetric is checked before symmetric to avoid a partial-match on
        "asymmetric" returning "Symmetric".  Falls back to inferring from side
        when no explicit keyword is present: Left/Right → Asymmetric,
        Universal → Symmetric.
        """
        if re.search(r"\basymmetric(?:al)?\b", text, re.I):
            return "Asymmetric"
        if re.search(r"\bsymmetric(?:al)?\b", text, re.I):
            return "Symmetric"
        if side in ("Left", "Right"):
            return "Asymmetric"
        if side == "Universal":
            return "Symmetric"
        return None

    @staticmethod
    def extract_fixation_surface(text: str) -> Optional[str]:
        """Return the primary fixation-surface feature keyword, or None.

        Checked in specificity order so "stemmed, keeled" → "Stemmed".
        Uses \\bstemmed\\b (not \\bstem\\b) to avoid false matches on
        generic phrases like "TKA System" or "Femoral Stem component".
        """
        if re.search(r"\bstemmed\b", text, re.I):
            return "Stemmed"
        if re.search(r"\bkeel(?:ed)?\b", text, re.I):
            return "Keeled"
        if re.search(r"\bpeg(?:ged)?\b", text, re.I):
            return "Pegged"
        if re.search(r"\bscrew\b", text, re.I):
            return "Screw"
        return None


class MDALLFormatter:
    """
    Formats catalogue numbers for MDALL lookup using CJRR-validated
    manufacturer-specific rules.  Retained for manual/targeted MDALL
    single-device lookups; bulk downloads use MDALLBulkDownloader directly.
    """

    @staticmethod
    def format_for_mdall(catalogue_num: str, manufacturer: str) -> str:
        if not catalogue_num or not manufacturer:
            return ""
        cat = re.sub(r"[^A-Z0-9]", "", str(catalogue_num).strip().upper())
        norm = MDALLFormatter._normalize_manufacturer(manufacturer)
        if norm == "Stryker/Osteonics/Howmedica":
            return MDALLFormatter._fmt_stryker(cat)
        if norm == "Zimmer/Biomet/Sulzer/Centerpulse":
            return MDALLFormatter._fmt_zimmer(cat)
        if norm == "DePuy/Finsbury/J & J":
            return MDALLFormatter._fmt_depuy(cat)
        return cat

    @staticmethod
    def _normalize_manufacturer(mfr: str) -> str:
        m = mfr.lower()
        if any(x in m for x in ["depuy", "finsbury", "j & j", "j&j", "johnson"]):
            return "DePuy/Finsbury/J & J"
        if any(x in m for x in ["zimmer", "biomet", "sulzer", "centerpulse"]):
            return "Zimmer/Biomet/Sulzer/Centerpulse"
        if any(x in m for x in ["stryker", "osteonics", "howmedica"]):
            return "Stryker/Osteonics/Howmedica"
        if "smith" in m and "nephew" in m:
            return "Smith & Nephew"
        if any(x in m for x in ["microport", "wright"]):
            return "MicroPort/Wright Medical"
        return mfr

    @staticmethod
    def _fmt_stryker(c: str) -> str:
        if re.search(r"\d[A-Z]\d", c):
            return re.sub(r"(\d)([A-Z])(\d)", r"\1-\2-\3", c)
        if re.search(r"\d{2}[A-Z0-9]*[A-Z]$", c):
            return c[:2] + "-" + c[2:]
        if len(c) == 7:
            return c[:2] + "-" + c[2] + "-" + c[3:]
        return c

    @staticmethod
    def _fmt_zimmer(c: str) -> str:
        if len(c) == 11:
            return f"{c[:2]}-{c[2:6]}-{c[6:9]}-{c[9:]}"
        if len(c) == 9:
            if c[4] == "0":
                return f"00-{c[:4]}-{c[4:7]}-{c[7:]}"
            return f"{c[:4]}-{c[4:6]}-{c[6:]}"
        return c

    @staticmethod
    def _fmt_depuy(c: str) -> str:
        if c[:3] == "178":
            return c
        if len(c) == 9:
            return f"{c[:4]}-{c[4:6]}-{c[6:]}"
        return c


class CSVSupplementer:
    """Supplements records with data from the NZ medical devices CSV."""

    RELEVANT_CATEGORIES = {
        "Femoral knee component",
        "Femoral knee components",
        "Femoral component",
        "Tibial baseplate",
        "Baseplates - tibial",
        "Tibial baseplates or trays",
        "Tibial Insert",
        "Tibial inserts or bearings or articular surfaces",
        "Patellar implant",
        "Patellar implants",
        "Patella implant",
        "Tibial extension stem",
        "Femoral knee extension stem",
        "Tibial or femoral stems",
        "Tibial extension stems",
        "Femoral knee extension stems",
        "Tibial stems",
        "Tibial components",
    }

    @staticmethod
    def load(csv_path: str) -> pd.DataFrame:
        logger.info(f"Loading supplemental CSV: {csv_path}")
        df = pd.read_csv(csv_path, encoding="utf-8-sig", dtype=str).fillna("")
        if "CategoryName3" in df.columns:
            df = df[df["CategoryName3"].isin(CSVSupplementer.RELEVANT_CATEGORIES)]
        logger.info(f"  {len(df):,} relevant rows loaded")
        return df

    @staticmethod
    def supplement(record: ImplantRecord, df: pd.DataFrame) -> None:
        """Fill missing fields on record from the CSV."""
        if not record.catalogue_num:
            return
        matches = df[df["SupplierProductCode"].astype(str) == str(record.catalogue_num)]
        if matches.empty and record.brand_name and record.company_name:
            matches = df[
                df["BrandName"].str.contains(record.brand_name, case=False, na=False)
                & df["SupplierName"].str.contains(
                    record.company_name, case=False, na=False
                )
            ]
        if matches.empty:
            return
        row = matches.iloc[0]
        if not record.size and row.get("DeviceName2", ""):
            record.size = row["DeviceName2"]
        extra = row.get("AdditionalInformation", "")
        if extra:
            record.notes = (record.notes + f" | CSV: {extra}").lstrip(" | ")
        if "Medical_Devices_CSV" not in record.data_source:
            record.data_source.append("Medical_Devices_CSV")


class FieldApplicabilityMatrix:
    """Marks fields as N/A or Missing based on component type."""

    APPLICABILITY: Dict[str, List[str]] = {
        "stability": ["Femoral", "Insert", "TKA System"],
        "bearing_type": ["Insert"],
        "fixation": ["Femoral", "Tibial", "TKA System"],
        "surface_treatment": ["Femoral", "Tibial"],
        "design_fixation_surface": ["Femoral", "Tibial", "Patellar"],
        "design_articular_surface": ["Femoral", "Insert"],
        "ap_length": ["Femoral", "Tibial"],
        "ml_width": ["Tibial"],
        "diameter": ["Patellar", "Femoral Stem", "Tibial Stem", "Stem"],
        "thickness": ["Insert"],
    }

    @staticmethod
    def mark(record: ImplantRecord) -> None:
        if not record.component_type:
            return
        for fname, applicable_types in FieldApplicabilityMatrix.APPLICABILITY.items():
            val = getattr(record, fname, None)
            if record.component_type not in applicable_types:
                record.field_status[fname] = FieldStatus.NOT_APPLICABLE.value
                setattr(record, fname, "N/A")
            elif val is None or val == "":
                record.field_status[fname] = FieldStatus.MISSING.value
                record.needs_review = True
            else:
                record.field_status[fname] = FieldStatus.PRESENT.value


# ===========================================================================
# GMDN term exclusions — non-implant records to drop after CJRR filter
# ===========================================================================

# Exact GMDN terms that are instruments, kits, hip implants, or data errors.
# Matched after stripping leading/trailing whitespace and quotes (CSV artefact).
_GMDN_EXCLUSIONS: frozenset = frozenset(
    {
        # Surgical instruments
        "Bone file",
        "Bone holding forceps",
        "Implant handling forceps",
        "Surgical implant handling forceps",
        "Orthopaedic bone calliper",
        "Orthopaedic broach",
        "Orthopaedic osteotome",
        "Orthopaedic implant aiming/guiding block",
        "Orthopaedic knee-balance analyser",
        "Optical surgical navigation device tracking system",
        # Kits / containers
        "Device sterilization/disinfection container",
        "Joint prosthesis implantation kit",
        "Orthopaedic surgical procedure kit",
        # Hip / acetabular implants (not knee)
        "Uncoated hip femur prosthesis",
        "Acetabular augmentation implant",
        "Acetabular shell",
        "Constrained polyethylene acetabular liner",
        "Non-constrained polyethylene acetabular liner",
        "Metallic acetabulum prosthesis",
        # Temporary / non-permanent devices
        "Femoral stem centralizer",
        "Orthopaedic cement spacer",
        # Data-quality artefacts
        "False",
        "True",
    }
)

# Keyword patterns to catch variants not in the exact list above
_GMDN_EXCLUSION_PATTERN = re.compile(
    r"forceps|screwdriver|mallet|drill\s*bit|punch|broach|osteotome"
    r"|calliper|caliper|acetabul|hip\s+femur|steriliz|navigation"
    r"|disinfect|splint|sidebar|clamp|cutter|file\b|hook\b",
    re.IGNORECASE,
)


# ===========================================================================
# Main pipeline
# ===========================================================================


class KneeImplantPipeline:
    """
    Orchestrates all pipeline stages.

    Stages can be entered at any point by supplying pre-built CSVs:
      - gudid_csv / mdall_csv  -> skip downloads
      - merged_csv             -> skip downloads + merge
      - cjrr_csv               -> skip downloads + merge + CJRR filter
    """

    def __init__(
        self,
        gudid_output_dir: str = "gudid_downloads",
        mdall_output_dir: str = "mdall_downloads",
        medical_devices_csv_path: Optional[str] = None,
        eifu_directory: Optional[str] = "implant_eIFUs",
        apply_cjrr_filter: bool = True,
        apply_gmdn_filter: bool = True,
        download_eifus: bool = False,  # auto-download missing eIFU PDFs
        download_eifus_headless: bool = False,
        skip_eifus: bool = False,  # disable ALL eIFU activity (download + extraction)
        enrich_gudid_descriptions: bool = True,  # fetch deviceDescription from GUDID API
        gudid_description_cache: str = "gudid_downloads/device_descriptions_cache.json",
        # --- short-circuit inputs (skip earlier stages) ---
        gudid_csv: Optional[str] = None,
        mdall_csv: Optional[str] = None,
        merged_csv: Optional[str] = None,
        cjrr_csv: Optional[str] = None,
        output_dir: str = "outputs",
    ):
        self.gudid_output_dir = gudid_output_dir
        self.mdall_output_dir = mdall_output_dir
        self.csv_path = medical_devices_csv_path
        self.eifu_dir = Path(eifu_directory) if eifu_directory else None
        self.apply_cjrr_filter = apply_cjrr_filter
        self.apply_gmdn_filter = apply_gmdn_filter
        self.download_eifus = download_eifus
        self.download_eifus_headless = download_eifus_headless
        self.skip_eifus = skip_eifus
        self.enrich_gudid_descriptions = enrich_gudid_descriptions
        self.gudid_description_cache = Path(gudid_description_cache)
        self.gudid_csv = gudid_csv
        self.mdall_csv = mdall_csv
        self.merged_csv = merged_csv
        self.cjrr_csv = cjrr_csv
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(exist_ok=True)

        self.records: List[ImplantRecord] = []
        self.csv_df: Optional[pd.DataFrame] = None

    # ------------------------------------------------------------------
    # Stage 1: GUDID download
    # ------------------------------------------------------------------

    def _get_gudid(self) -> pd.DataFrame:
        """Return GUDID DataFrame — download or load from existing file."""
        if self.gudid_csv:
            path = (
                _merger.load_latest(self.gudid_csv)
                if "*" in self.gudid_csv
                else pd.read_csv(self.gudid_csv, dtype=str).fillna("")
            )
            if isinstance(path, pd.DataFrame):
                df = path
            else:
                df = path
            logger.info(f"Loaded GUDID from file: {len(df):,} rows")
            return df

        # Try to find the latest existing download
        existing = sorted(
            Path(self.gudid_output_dir).glob("knee_devices_bulk_*.csv"),
            key=lambda p: p.stat().st_mtime,
        )
        if existing:
            logger.info(f"Using existing GUDID file: {existing[-1]}")
            df = pd.read_csv(existing[-1], dtype=str).fillna("")
            logger.info(f"  {len(df):,} rows")
            return df

        # Fresh download
        logger.info("Downloading GUDID data (bulk product code download)...")
        downloader = ProductCodeBulkDownloader(output_dir=self.gudid_output_dir)
        df = downloader.download_all_knee_codes()
        if self.apply_gmdn_filter and not df.empty:
            df = downloader.apply_gmdn_filtering(df)
        return df

    # ------------------------------------------------------------------
    # Stage 2: MDALL download
    # ------------------------------------------------------------------

    def _get_mdall(self) -> pd.DataFrame:
        """Return MDALL DataFrame — download or load from existing file."""
        if self.mdall_csv:
            df = pd.read_csv(self.mdall_csv, dtype=str).fillna("")
            logger.info(f"Loaded MDALL from file: {len(df):,} rows")
            return df

        existing = sorted(
            Path(self.mdall_output_dir).glob("mdall_knee_devices_*.csv"),
            key=lambda p: p.stat().st_mtime,
        )
        if existing:
            logger.info(f"Using existing MDALL file: {existing[-1]}")
            df = pd.read_csv(existing[-1], dtype=str).fillna("")
            logger.info(f"  {len(df):,} rows")
            return df

        logger.info("Downloading MDALL data...")
        downloader = MDALLBulkDownloader(output_dir=self.mdall_output_dir)
        df = downloader.download_all(fetch_identifiers=True)
        return df

    # ------------------------------------------------------------------
    # Stage 3: Merge
    # ------------------------------------------------------------------

    def _get_merged(
        self, gudid_df: pd.DataFrame, mdall_df: pd.DataFrame
    ) -> pd.DataFrame:
        """Merge GUDID and MDALL DataFrames using the two-pass catalogue match."""
        if self.merged_csv:
            df = pd.read_csv(self.merged_csv, dtype=str).fillna("")
            logger.info(f"Loaded merged CSV: {len(df):,} rows")
            return df

        # If both inputs are empty (--cjrr path skips download+merge),
        # return an empty DataFrame — _apply_cjrr will load the CJRR CSV.
        if gudid_df.empty and mdall_df.empty:
            existing = sorted(
                Path(".").glob("merged_gudid_mdall*.csv"),
                key=lambda p: p.stat().st_mtime,
            )
            if existing:
                logger.info(f"Using existing merged file: {existing[-1]}")
                return pd.read_csv(existing[-1], dtype=str).fillna("")
            return pd.DataFrame()

        logger.info("Merging GUDID and MDALL...")
        merged = _merger.merge(gudid_df, mdall_df)
        _merger.print_summary(merged)

        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        out = self.output_dir / f"merged_gudid_mdall_{ts}.csv"
        merged.to_csv(out, index=False, encoding="utf-8")
        logger.info(f"[OK] Merged CSV saved: {out}")
        return merged

    # ------------------------------------------------------------------
    # Stage 4: CJRR filter
    # ------------------------------------------------------------------

    def _filter_gmdn_exclusions(self, df: pd.DataFrame) -> pd.DataFrame:
        """Drop records whose GMDN term is an instrument, kit, hip implant, etc."""
        if "GMDN_TERMS" not in df.columns:
            return df
        # Normalise: strip whitespace + stray leading/trailing quotes (CSV artefact)
        normalized = df["GMDN_TERMS"].astype(str).str.strip().str.strip('"').str.strip()
        exact_mask = normalized.isin(_GMDN_EXCLUSIONS)
        keyword_mask = normalized.str.contains(_GMDN_EXCLUSION_PATTERN, na=False)
        excluded = exact_mask | keyword_mask
        n = excluded.sum()
        if n:
            logger.info(
                "GMDN exclusion filter: removing %d records with non-implant GMDN terms",
                n,
            )
            for term, cnt in df.loc[excluded, "GMDN_TERMS"].value_counts().items():
                logger.info("  excluded %3d  %r", cnt, term)
        return df[~excluded].reset_index(drop=True)

    def _apply_cjrr(self, df: pd.DataFrame) -> pd.DataFrame:
        """Filter to CJRR manufacturers and exclude partial knees."""
        if self.cjrr_csv:
            filtered = pd.read_csv(self.cjrr_csv, dtype=str).fillna("")
            logger.info(f"Loaded CJRR-filtered CSV: {len(filtered):,} rows")
            return filtered

        if not self.apply_cjrr_filter:
            return df

        logger.info("Applying CJRR manufacturer filter...")
        filtered = _cjrr_filter.filter_cjrr(df)
        _cjrr_filter.print_summary(df, filtered)

        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        out = self.output_dir / f"cjrr_filtered_{ts}.csv"
        filtered.to_csv(out, index=False, encoding="utf-8")
        logger.info(f"[OK] CJRR-filtered CSV saved: {out}")
        return filtered

    # ------------------------------------------------------------------
    # Stage 5: Parse DataFrame rows -> ImplantRecord
    # ------------------------------------------------------------------

    @staticmethod
    def _row_to_record(row: pd.Series) -> ImplantRecord:
        """Convert a merged/filtered DataFrame row into an ImplantRecord."""
        r = ImplantRecord()
        r.version_model_number = row.get("VERSION_MODEL_NUMBER", "") or None
        r.catalogue_num = r.version_model_number
        r.brand_name = row.get("BRAND_NAME", "") or None
        r.mdall_brand_name = row.get("MDALL_BRAND_NAME", "") or None
        r.company_name = row.get("COMPANY_NAME", "") or None
        r.manufacturer = row.get("CJRR_MANUFACTURER", "") or r.company_name
        r.device_id = row.get("DEVICE_ID", "") or None
        r.device_description = row.get("DEVICE_DESCRIPTION", "") or None
        r.device_sizes_text = row.get("DEVICE_SIZES_TEXT", "") or None
        r.gmdn_name = row.get("GMDN_TERMS", "") or None
        r.product_code = row.get("product_code", "") or None
        r.device_name = row.get("product_code_description", "") or None
        r.mdall_licence_number = row.get("LICENCE_NUMBER", "") or None
        r.mdall_device_id = row.get("MDALL_DEVICE_ID", "") or None
        r.match_type = row.get("MATCH_TYPE", "") or None
        r.source_db = row.get("SOURCE", "") or None

        # Populate data_source list
        src = row.get("SOURCE", "")
        if "GUDID" in src or src == "BOTH":
            r.data_source.append("FDA_GUDID")
        if "MDALL" in src or src == "BOTH":
            r.data_source.append("MDALL")

        # Quick text-based material hints from GMDN
        if r.gmdn_name:
            metal, polymer, antioxidant = FieldExtractor.extract_material(r.gmdn_name)
            r.metal_material = metal
            r.poly_material = polymer
            r.antioxidant = antioxidant

        notes_parts = []
        mri = row.get("MRI_SAFETY_STATUS", "")
        if mri:
            notes_parts.append(f"MRI: {mri}")
        lic_status = row.get("LICENCE_STATUS", "")
        if lic_status:
            notes_parts.append(f"MDALL status: {lic_status}")
        r.notes = " | ".join(notes_parts)

        return r

    # ------------------------------------------------------------------
    # Stage 4c: GUDID device description enrichment
    # ------------------------------------------------------------------

    def _maybe_enrich_gudid_descriptions(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        Fetch deviceDescription + deviceSizes from the GUDID REST API for
        records that have a DEVICE_ID.  Results are cached locally so
        subsequent runs only fetch new DIs.  When enrich_gudid_descriptions=False
        (--no-gudid-api), no API calls are made but cached results are still applied.
        """
        if not self.enrich_gudid_descriptions:
            logger.info(
                "GUDID description enrichment: API disabled, applying cached results only"
            )
            return _enrich_gudid_descriptions(
                df,
                cache_path=self.gudid_description_cache,
                dry_run=True,
            )
        return _enrich_gudid_descriptions(
            df,
            cache_path=self.gudid_description_cache,
        )

    # ------------------------------------------------------------------
    # Stage 5b: Download missing eIFU PDFs (optional)
    # ------------------------------------------------------------------

    def _maybe_download_eifus(self, base_df: pd.DataFrame) -> None:
        """
        If download_eifus=True, trigger EIFUDownloadOrchestrator to fetch
        any eIFU PDFs for systems present in base_df that are not yet
        in eifu_dir.  Runs before _build_eifu_lookup so the freshly
        downloaded PDFs are available immediately.
        """
        if not self.download_eifus or not self.eifu_dir:
            return
        logger.info("Checking for missing eIFU PDFs...")
        orch = EIFUDownloadOrchestrator(
            eifu_dir=str(self.eifu_dir),
            headless=self.download_eifus_headless,
        )
        results = orch.download_missing_for_dataframe(base_df)
        if results:
            orch.print_summary(results)

    # ------------------------------------------------------------------
    # Stage 6: eIFU material enrichment
    # ------------------------------------------------------------------

    def _build_eifu_lookup(self) -> Dict[str, Dict]:
        """
        Scan the eIFU directory, extract materials from each PDF, and build
        a flat lookup with two key strategies:

          (a) Normalised catalogue family prefix — Stryker catalogue-number match
              e.g. "5512FXXX" or "5512F" for family "5512-F-XXX"

          (b) "sys:{manufacturer}:{system}:{component_type}" — Zimmer/DePuy match
              e.g. "sys:zimmer:persona:femoral"

        Values are eIFU material record dicts as returned by extract_materials().
        """
        lookup: Dict[str, Dict] = {}

        if not self.eifu_dir or not self.eifu_dir.exists():
            return lookup

        pdf_files = list(self.eifu_dir.glob("*.pdf"))
        if not pdf_files:
            logger.info("No eIFU PDFs found in %s", self.eifu_dir)
            return lookup

        logger.info(f"Extracting materials from {len(pdf_files)} eIFU PDF(s)...")

        for pdf in pdf_files:
            try:
                records = _extract_eifu_materials(pdf)
                n_before = len(lookup)

                variant = _detect_eifu_variant(pdf)

                for rec in records:
                    # (a) Catalogue-family keys (Stryker)
                    cat_fam = rec.get("catalogue_family")
                    if cat_fam:
                        for k in _eifu_keys(cat_fam):
                            lookup.setdefault(k, rec)

                    # (b) System-level key (all manufacturers)
                    mfr = rec.get("manufacturer", "").lower()
                    sys_ = rec.get("system", "").lower().replace(" ", "_")
                    comp = rec.get("component_type", "").lower()
                    if mfr and sys_ and comp:
                        # Variant-specific key (e.g. attune_cementless) takes
                        # priority; bare key acts as fallback for unmatched fixation
                        if variant:
                            lookup.setdefault(f"sys:{mfr}:{sys_}:{variant}:{comp}", rec)
                        lookup.setdefault(f"sys:{mfr}:{sys_}:{comp}", rec)

                logger.info(
                    "  %s: %d records, %d new lookup keys",
                    pdf.name,
                    len(records),
                    len(lookup) - n_before,
                )
            except Exception as exc:
                logger.warning(f"  {pdf.name}: extraction failed — {exc}")

        logger.info(f"eIFU lookup: {len(lookup)} total keys")
        return lookup

    @staticmethod
    def _enrich_from_eifu(record: ImplantRecord, eifu_lookup: Dict[str, Dict]) -> None:
        """Populate material fields on record from the eIFU lookup."""
        if not eifu_lookup:
            return

        mat: Optional[Dict] = None

        # (a) Catalogue-number prefix match (works for Stryker)
        if record.catalogue_num:
            mat = _find_eifu_match(record.catalogue_num, eifu_lookup)

        # (b) System + component_type match (Zimmer / DePuy)
        if not mat:
            mapping = _brand_to_system(record.brand_name or "")
            if mapping:
                _, _, system = mapping
                mfr_short = (
                    "stryker"
                    if "stryker" in (record.manufacturer or "").lower()
                    else (
                        "zimmer"
                        if "zimmer" in (record.manufacturer or "").lower()
                        else (
                            "depuy"
                            if "depuy" in (record.manufacturer or "").lower()
                            else ""
                        )
                    )
                )
                comp = (record.component_type or "").lower()
                if mfr_short and system and comp:
                    sys_lower = system.lower()
                    # Derive variant from known fixation so we prefer the matching PDF
                    fixation = (record.fixation or "").lower()
                    variant = (
                        "cementless"
                        if fixation in ("cementless", "uncemented")
                        else (
                            "cemented"
                            if fixation == "cemented"
                            else (
                                "mobile"
                                if fixation == "mobile"
                                else "fixed" if fixation == "fixed" else ""
                            )
                        )
                    )
                    # Try variant-specific key first, then bare system key
                    if variant:
                        mat = eifu_lookup.get(
                            f"sys:{mfr_short}:{sys_lower}:{variant}:{comp}"
                        )
                    if not mat:
                        mat = eifu_lookup.get(f"sys:{mfr_short}:{sys_lower}:{comp}")

        if not mat:
            return

        mat_name = mat.get("material_name", "").lower()
        if not record.metal_material and mat_name:
            if "cobalt" in mat_name or "co-cr" in mat_name or "cocr" in mat_name:
                record.metal_material = "CoCr"
            elif "titanium" in mat_name or "ti-6al" in mat_name:
                record.metal_material = "Titanium"
            elif "oxidized" in mat_name or "zirconium" in mat_name:
                record.metal_material = "OxZr"
        if not record.poly_material and mat_name:
            if "uhmwpe" in mat_name or "polyethylene" in mat_name:
                record.poly_material = (
                    "UHMWPE+VitE"
                    if ("vitamin" in mat_name or "vehxpe" in mat_name)
                    else "UHMWPE"
                )

        alloy_std = mat.get("alloy_standard")
        if alloy_std and not record.material_standard:
            record.material_standard = alloy_std

        # Surface treatment from eIFU material name
        # (PA = Peri-Apatite → HA; Hydroxyapatite → HA; TiN → TiN)
        if not record.surface_treatment and mat_name:
            if re.search(r"hydroxyapatite|peri.?apatite", mat_name):
                record.surface_treatment = "HA"
            elif re.search(r"\btin\b|titanium\s+nitride", mat_name):
                record.surface_treatment = "TiN"
            elif re.search(r"titanium\s+plasma|plasma\s+spray|\btps\b", mat_name):
                record.surface_treatment = "TPS"

        if "eIFU" not in record.data_source:
            record.data_source.append("eIFU")

    # ------------------------------------------------------------------
    # Stage 7: General enrichment (classify + extract fields)
    # ------------------------------------------------------------------

    def _enrich(self, record: ImplantRecord, eifu_lookup: Dict[str, Dict]) -> None:
        """Classify component type, extract design fields, enrich from eIFU."""
        # Most-specific fields first so re.search() finds device-specific text
        # before generic category labels.  device_name (FDA product-code
        # description, e.g. "Knee Joint Femoral Component") and gmdn_name
        # (GMDN category, e.g. "Semi-constrained knee prosthesis") are placed
        # last because they describe the device class, not the specific implant.
        search_text = " ".join(
            v
            for v in [
                record.device_description,
                record.device_sizes_text,  # GUDID size annotations: "Side: LEFT | Type: CR | Size: 13"
                record.mdall_brand_name,
                record.brand_name,
                record.version_model_number,
                record.notes,
                record.device_name,
                record.gmdn_name,
            ]
            if isinstance(v, str) and v
        )

        # Component type
        if not record.component_type:
            record.component_type = ComponentTypeClassifier.classify(
                gmdn_name=record.gmdn_name or "",
                device_description=record.device_description or "",
            )

        # Design fields
        # For stability: GMDN terms all say "semi-constrained" (GMDN category for
        # knee inserts) which overrides actual CR/PS info in brand names.
        # Use brand/description text first; only fall through if still unset.
        # Stability search excludes device_name (FDA product code description) and
        # gmdn_name because both generically say "semi-constrained" for all knee
        # inserts regardless of actual CR/PS/CCK design.  Use only fields that
        # carry device-specific information.
        brand_text = " ".join(
            v
            for v in [
                record.device_description,
                record.mdall_brand_name,
                record.brand_name,
                record.version_model_number,
                record.notes,
            ]
            if v and isinstance(v, str)
        )
        if not record.fixation:
            record.fixation = FieldExtractor.extract_fixation(search_text)
        if not record.stability:
            record.stability = FieldExtractor.extract_stability(brand_text)
        if not record.bearing_type:
            record.bearing_type = FieldExtractor.extract_bearing_type(search_text)
        if not record.surface_treatment:
            record.surface_treatment = FieldExtractor.extract_surface_treatment(
                search_text
            )
        if not record.side:
            record.side = FieldExtractor.extract_side(search_text)
        if not record.design_articular_surface:
            record.design_articular_surface = FieldExtractor.extract_articular_surface(
                search_text, side=record.side
            )
        if not record.design_fixation_surface:
            record.design_fixation_surface = FieldExtractor.extract_fixation_surface(
                search_text
            )
        if not record.size:
            record.size = FieldExtractor.extract_size(search_text)
        if not record.thickness:
            t_str = FieldExtractor.extract_thickness(search_text)
            if t_str:
                try:
                    record.thickness = float(t_str)
                except (ValueError, TypeError):
                    pass

        # Structured dimensions from device_sizes_text (GUDID API)
        if isinstance(record.device_sizes_text, str) and record.device_sizes_text:
            dims = FieldExtractor.extract_dimensions_from_sizes_text(
                record.device_sizes_text
            )
            if not record.ap_length and dims["ap_length"]:
                record.ap_length = dims["ap_length"]
            if not record.ml_width and dims["ml_width"]:
                record.ml_width = dims["ml_width"]
            if not record.thickness and dims["thickness"]:
                record.thickness = dims["thickness"]
            if not record.diameter and dims["diameter"]:
                record.diameter = dims["diameter"]

        # Materials from text — re-run if any field is unset, or if poly_material is
        # only the generic GMDN-level "UHMWPE" and device-specific text may carry a
        # more specific crosslinked type (XLPE, HXLPE, HXLPE+VitE, etc.).
        if (
            not record.metal_material
            or not record.poly_material
            or record.poly_material == "UHMWPE"
        ):
            metal, polymer, antioxidant = FieldExtractor.extract_material(search_text)
            if not record.metal_material:
                record.metal_material = metal
            if not record.poly_material or (
                record.poly_material == "UHMWPE" and polymer and polymer != "UHMWPE"
            ):
                record.poly_material = polymer
            if not record.antioxidant:
                record.antioxidant = antioxidant

        # eIFU material enrichment
        self._enrich_from_eifu(record, eifu_lookup)

        # Catalogue-number size extraction (fallback for fields still unset)
        cat_fields = _cat_size(
            catalogue_num=record.catalogue_num or "",
            manufacturer=record.manufacturer or record.company_name or "",
            component_type=record.component_type or "",
            brand_name=record.brand_name or "",
        )
        if not record.size and cat_fields["size"]:
            record.size = cat_fields["size"]
        if not record.side and cat_fields["side"]:
            record.side = cat_fields["side"]
        if not record.thickness and cat_fields["thickness"]:
            try:
                record.thickness = float(cat_fields["thickness"])
            except (ValueError, TypeError):
                pass
        if not record.stability and cat_fields["stability"]:
            record.stability = cat_fields["stability"]
        if not record.poly_material and cat_fields.get("poly_material"):
            record.poly_material = cat_fields["poly_material"]
        if not record.antioxidant and cat_fields.get("antioxidant"):
            record.antioxidant = cat_fields["antioxidant"]
        if not record.diameter and cat_fields.get("diameter"):
            try:
                record.diameter = float(cat_fields["diameter"])
            except (ValueError, TypeError):
                pass

        # Confidence score: fraction of non-empty non-meta fields
        total = 0
        filled = 0
        skip = {
            "field_status",
            "data_source",
            "last_updated",
            "notes",
            "needs_review",
            "confidence_score",
        }
        for k, v in asdict(record).items():
            if k in skip:
                continue
            total += 1
            if v is not None and v != "" and v != []:
                filled += 1
        record.confidence_score = filled / total if total else 0.0

        # Field applicability matrix
        FieldApplicabilityMatrix.mark(record)

    # ------------------------------------------------------------------
    # Export
    # ------------------------------------------------------------------

    # Fields where we take the first non-empty value across all rows sharing
    # a catalogue number.  Order matters: we try primary first, then fill from
    # secondary rows only when the primary is empty.
    def _to_dataframe(self) -> pd.DataFrame:
        rows = [asdict(r) for r in self.records]
        df = pd.DataFrame(rows)
        if df.empty:
            return df

        df["data_source"] = df["data_source"].apply(
            lambda x: ", ".join(x) if isinstance(x, list) else x
        )
        df["field_status"] = df["field_status"].apply(json.dumps)

        # Deduplicate on catalogue_num: a catalogue number may appear under
        # multiple FDA product codes (e.g. KWH + NJL).  Each row has already
        # been individually enriched, so secondary rows may carry fields the
        # primary lacks (e.g. bearing_type from the NJL mobile-bearing code).
        # Strategy:
        #   1. Sort by source quality (BOTH > GUDID_ONLY > MDALL_ONLY) then
        #      confidence_score descending so iloc[0] is always the best row.
        #   2. Coalesce empty fields in the primary from secondary rows.
        #   3. Collect all product codes into a pipe-separated string.
        if "catalogue_num" in df.columns:
            n_before = len(df)
            source_order = {"BOTH": 0, "GUDID_ONLY": 1, "MDALL_ONLY": 2}
            df["_src_rank"] = df["source_db"].map(source_order).fillna(9)
            df = df.sort_values(
                ["_src_rank", "confidence_score"], ascending=[True, False]
            )
            df = (
                df.groupby("catalogue_num", sort=False, group_keys=False)
                .apply(_merge_catalogue_group)
                .drop(columns=["_src_rank"])
                .reset_index(drop=True)
            )
            n_removed = n_before - len(df)
            if n_removed:
                logger.info(
                    "Merged %d duplicate catalogue number(s) into %d unique records"
                    " (fields coalesced, all product codes preserved)",
                    n_removed,
                    len(df),
                )

        return df

    # ------------------------------------------------------------------
    # Post-dedup review flag recomputation
    # ------------------------------------------------------------------

    def _recompute_review_flags(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        Recompute needs_review and field_status for every row after deduplication.

        _merge_catalogue_group() coalesces empty fields from secondary rows into
        the primary row, but the primary's needs_review and field_status were set
        before that merge.  A field marked MISSING may now be filled; this method
        re-evaluates every applicable field against the actual values in the
        merged DataFrame so the final CSV and review reports are accurate.
        """
        APPLICABILITY = FieldApplicabilityMatrix.APPLICABILITY

        def _recompute_row(row: pd.Series) -> pd.Series:
            comp = row.get("component_type") or ""
            try:
                fs = (
                    json.loads(row["field_status"])
                    if isinstance(row["field_status"], str)
                    else {}
                )
            except Exception:
                fs = {}

            if not comp:
                # No component type — can't evaluate applicability; leave unchanged
                return pd.Series(
                    {
                        "needs_review": row.get("needs_review", False),
                        "field_status": row["field_status"],
                    }
                )

            needs_review = False
            for fname, applicable_types in APPLICABILITY.items():
                if fname not in df.columns:
                    continue
                val = row.get(fname)
                if comp not in applicable_types:
                    fs[fname] = FieldStatus.NOT_APPLICABLE.value
                elif _is_empty(val) or str(val).strip() in ("N/A", "n/a"):
                    fs[fname] = FieldStatus.MISSING.value
                    needs_review = True
                else:
                    fs[fname] = FieldStatus.PRESENT.value

            return pd.Series(
                {"needs_review": needs_review, "field_status": json.dumps(fs)}
            )

        recomputed = df.apply(_recompute_row, axis=1)
        df = df.copy()
        df["needs_review"] = recomputed["needs_review"]
        df["field_status"] = recomputed["field_status"]

        n_flagged = int(df["needs_review"].sum())
        logger.info(
            "Post-dedup review flags recomputed: %d / %d records need review",
            n_flagged,
            len(df),
        )
        return df

    # ------------------------------------------------------------------
    # Summary
    # ------------------------------------------------------------------

    def _print_summary(self, df: pd.DataFrame) -> None:
        logger.info("\n" + "=" * 70)
        logger.info("PIPELINE SUMMARY")
        logger.info("=" * 70)
        logger.info(f"Total records:           {len(df):>8,}")
        if df.empty:
            return
        logger.info(f"Records needing review:  {df['needs_review'].sum():>8,}")
        logger.info(f"Avg confidence score:    {df['confidence_score'].mean():>8.2f}")

        if "component_type" in df.columns:
            logger.info("\nComponent types:")
            for ct, n in df["component_type"].value_counts().items():
                logger.info(f"  {ct:<35} {n:>7,}")

        if "manufacturer" in df.columns:
            logger.info("\nTop manufacturers (CJRR canonical):")
            for mfr, n in df["manufacturer"].value_counts().head(15).items():
                logger.info(f"  {mfr:<40} {n:>7,}")

        if "source_db" in df.columns:
            logger.info("\nData source:")
            for src, n in df["source_db"].value_counts().items():
                logger.info(f"  {src:<25} {n:>7,}")
        logger.info("=" * 70)

    # ------------------------------------------------------------------
    # Main entry point
    # ------------------------------------------------------------------

    def run(self) -> pd.DataFrame:
        logger.info("=" * 70)
        logger.info("KNEE IMPLANT REFERENCE LIBRARY PIPELINE  v3")
        logger.info("=" * 70)

        # -- Stage 1 + 2: download (may be skipped) ----------------------
        if self.cjrr_csv or self.merged_csv:
            # Jump straight to the later stage
            gudid_df = pd.DataFrame()
            mdall_df = pd.DataFrame()
        else:
            gudid_df = self._get_gudid()
            logger.info(f"GUDID: {len(gudid_df):,} rows")
            # Filter non-implant GMDN terms from GUDID *before* merging with
            # MDALL so instruments / kits / hip devices are never matched
            gudid_df = self._filter_gmdn_exclusions(gudid_df)
            logger.info(f"GUDID after GMDN exclusion: {len(gudid_df):,} rows")
            mdall_df = self._get_mdall()
            logger.info(f"MDALL: {len(mdall_df):,} rows")

        # -- Stage 3: merge ----------------------------------------------
        merged_df = self._get_merged(gudid_df, mdall_df)
        logger.info(f"Merged: {len(merged_df):,} rows")

        # -- Stage 4: CJRR filter ----------------------------------------
        base_df = self._apply_cjrr(merged_df)
        logger.info(f"After CJRR filter: {len(base_df):,} rows")
        # Safety net: also apply GMDN exclusion when entering from --merged
        # or --cjrr (GUDID step was skipped so filter has not run yet)
        if self.merged_csv or self.cjrr_csv:
            base_df = self._filter_gmdn_exclusions(base_df)
            logger.info(f"After GMDN exclusion filter: {len(base_df):,} rows")

        # -- Stage 4c: GUDID device description enrichment ---------------
        base_df = self._maybe_enrich_gudid_descriptions(base_df)

        # -- Stage 5a: Download missing eIFUs (optional) -----------------
        if not self.skip_eifus:
            self._maybe_download_eifus(base_df)
        else:
            logger.info("eIFU download skipped (skip_eifus=True)")

        # -- Stage 5b: CSV supplement ------------------------------------
        if self.csv_path and Path(self.csv_path).exists():
            self.csv_df = CSVSupplementer.load(self.csv_path)

        # -- Stage 6: eIFU lookup ----------------------------------------
        if not self.skip_eifus:
            eifu_lookup = self._build_eifu_lookup()
        else:
            eifu_lookup = {}
            logger.info("eIFU material extraction skipped (skip_eifus=True)")

        # -- Stage 7: parse rows -> ImplantRecord + enrich ---------------
        logger.info(f"Building and enriching {len(base_df):,} ImplantRecords...")
        self.records = []
        for _, row in base_df.iterrows():
            record = self._row_to_record(row)
            if self.csv_df is not None:
                CSVSupplementer.supplement(record, self.csv_df)
            self._enrich(record, eifu_lookup)
            self.records.append(record)

        # -- Stage 8: export ---------------------------------------------
        df = self._to_dataframe()
        df = self._recompute_review_flags(df)
        self._print_summary(df)

        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        main_out = self.output_dir / f"knee_implant_library_{ts}.csv"
        df.to_csv(main_out, index=False, encoding="utf-8")
        logger.info(f"\n[OK] Full library saved: {main_out}  ({len(df):,} records)")

        # Excel export (human-readable, prevents date auto-conversion)
        xlsx_out = self.output_dir / f"knee_implant_library_{ts}.xlsx"
        df.to_excel(xlsx_out, index=False)

        wb = load_workbook(xlsx_out)
        ws = wb.active
        col_letter = next(cell.column_letter for cell in ws[1] if cell.value == "size")
        for cell in ws[col_letter][1:]:  # skip header
            cell.number_format = FORMAT_TEXT
            cell.value = str(cell.value)
        wb.save(xlsx_out)
        logger.info(f"[OK] Excel library saved: {xlsx_out}  ({len(df):,} records)")

        # Needs-review subset + detailed review report
        self._generate_review_report(df, ts)

        return df

    # ------------------------------------------------------------------
    # Review report
    # ------------------------------------------------------------------

    def _generate_review_report(self, df: pd.DataFrame, ts: str) -> None:
        """
        Build and save two complementary review outputs:

          needs_review_{ts}.csv
            One row per manufacturer/brand/component_type group that has at
            least one record needing review.  Editable field columns are
            pre-populated with the current consensus value (if consistent
            across the group) or left blank (mixed / missing).  Fill in the
            blanks and feed the file back via apply_review.py.

            Columns: manufacturer, brand_name, component_type, n_records,
                     n_needing_review, missing_fields, has_eifu_coverage,
                     <editable fields ...>, catalogue_nums

          review_report_{ts}.csv
            Aggregated summary of ALL groups (not just those needing review)
            for reference.
        """
        if df.empty:
            return

        from collections import Counter

        # Fields the user can fill in — order determines column order in CSV
        _FILLABLE = [
            "stability",
            "fixation",
            "bearing_type",
            "metal_material",
            "poly_material",
            "antioxidant",
            "surface_treatment",
            "material_standard",
            "design_fixation_surface",
            "design_articular_surface",
        ]

        def _consensus(series: pd.Series) -> str:
            """Single consistent non-null/non-NA value, else blank."""
            vals = (
                series.replace({"": None, "N/A": None, "n/a": None}).dropna().unique()
            )
            return str(vals[0]) if len(vals) == 1 else ""

        def _missing_fields(field_status_series: pd.Series) -> List[str]:
            missing: set = set()
            for fs_raw in field_status_series:
                try:
                    fs = (
                        json.loads(fs_raw)
                        if isinstance(fs_raw, str)
                        else (fs_raw or {})
                    )
                except Exception:
                    fs = {}
                missing.update(
                    k for k, v in fs.items() if v == FieldStatus.MISSING.value
                )
            return sorted(missing)

        # ----------------------------------------------------------------
        # 1. Grouped fillable needs-review
        # ----------------------------------------------------------------
        review_rows = []
        for (mfr, brand, comp), grp in df.groupby(
            ["manufacturer", "brand_name", "component_type"], dropna=False
        ):
            if not grp["needs_review"].any():
                continue

            has_eifu = grp["data_source"].str.contains("eIFU", na=False).any()
            missing = _missing_fields(grp["field_status"])

            row: Dict = {
                "manufacturer": mfr,
                "brand_name": brand,
                "component_type": comp,
                "n_records": len(grp),
                "n_needing_review": int(grp["needs_review"].sum()),
                "missing_fields": ", ".join(missing),
                "has_eifu_coverage": bool(has_eifu),
            }
            for fname in _FILLABLE:
                row[fname] = _consensus(grp[fname]) if fname in grp.columns else ""
            row["catalogue_nums"] = " | ".join(
                grp["catalogue_num"].dropna().unique().tolist()
            )
            review_rows.append(row)

        review_df = pd.DataFrame(review_rows).sort_values(
            ["manufacturer", "brand_name", "component_type"]
        )
        if not review_df.empty:
            rev_out = self.output_dir / f"needs_review_{ts}.csv"
            review_df.to_csv(rev_out, index=False, encoding="utf-8")
            logger.info(
                "[OK] Needs-review (grouped, fillable) saved: %s  (%d groups)",
                rev_out,
                len(review_df),
            )

        # ----------------------------------------------------------------
        # 2. Full aggregated summary (all groups, for reference)
        # ----------------------------------------------------------------
        agg_rows = []
        for (mfr, brand, comp), grp in df.groupby(
            ["manufacturer", "brand_name", "component_type"], dropna=False
        ):
            missing = _missing_fields(grp["field_status"])
            has_eifu = grp["data_source"].str.contains("eIFU", na=False).any()
            agg_rows.append(
                {
                    "manufacturer": mfr,
                    "brand_name": brand,
                    "component_type": comp,
                    "n_records": len(grp),
                    "n_needing_review": int(grp["needs_review"].sum()),
                    "has_eifu_coverage": bool(has_eifu),
                    "pct_metal_filled": f"{(grp['metal_material'].ne('') & grp['metal_material'].notna()).mean()*100:.0f}%",
                    "pct_poly_filled": f"{(grp['poly_material'].ne('') & grp['poly_material'].notna()).mean()*100:.0f}%",
                    "missing_fields": ", ".join(missing),
                    "sample_cat_nums": " | ".join(
                        grp["catalogue_num"].dropna().unique()[:5].tolist()
                    ),
                }
            )

        agg_df = pd.DataFrame(agg_rows).sort_values(
            ["manufacturer", "brand_name", "component_type"]
        )
        rpt_out = self.output_dir / f"review_report_{ts}.csv"
        agg_df.to_csv(rpt_out, index=False, encoding="utf-8")
        logger.info(
            "[OK] Review report saved: %s  (%d system/component rows)",
            rpt_out,
            len(agg_df),
        )

        # ----------------------------------------------------------------
        # 3. Console summary
        # ----------------------------------------------------------------
        needs_manual = agg_df[agg_df["n_needing_review"] > 0]
        no_eifu = agg_df[~agg_df["has_eifu_coverage"]]

        logger.info("\n" + "=" * 70)
        logger.info("MANUAL ENTRY REQUIRED — SUMMARY")
        logger.info("=" * 70)
        logger.info("  %d system/component groups need attention", len(needs_manual))
        logger.info(
            "  %d individual records flagged", int(agg_df["n_needing_review"].sum())
        )

        if not no_eifu.empty:
            logger.info("\n  Groups with NO eIFU coverage (%d):", len(no_eifu))
            for _, r in no_eifu.iterrows():
                logger.info(
                    "    %-35s  %-30s  %-20s  %d records",
                    r["manufacturer"],
                    r["brand_name"],
                    r["component_type"],
                    r["n_records"],
                )

        if not needs_manual.empty:
            logger.info("\n  Missing-field breakdown (fields needing manual entry):")
            all_mf: List[str] = []
            for mf in needs_manual["missing_fields"]:
                all_mf.extend(f.strip() for f in mf.split(",") if f.strip())
            for fname, cnt in Counter(all_mf).most_common():
                logger.info("    %-35s  %d groups", fname, cnt)

        logger.info("=" * 70)


# ===========================================================================
# eIFU catalogue matching helpers (module-level, used by pipeline)
# ===========================================================================


def _norm_cat(value: str) -> str:
    """Normalise catalogue number: uppercase, strip hyphens + spaces."""
    return re.sub(r"[-\s]", "", value.upper())


def _eifu_keys(catalogue_family: str) -> List[str]:
    """
    Return lookup keys for a catalogue family.

    e.g. "5512-F-XXX" -> ["5512FXXX", "5512F"]
         "5512-F-101" -> ["5512F101", "5512F"]
    """
    norm = _norm_cat(catalogue_family)
    keys = [norm]
    # Also add prefix without trailing digits or XXX suffix
    prefix = re.sub(r"(?:XXX|\d{3})$", "", norm)
    if prefix and prefix != norm:
        keys.append(prefix)
    return keys


def _find_eifu_match(catalogue_num: str, lookup: Dict[str, Dict]) -> Optional[Dict]:
    """Match a catalogue number against the eIFU lookup (exact, then prefix).
    Only scans catalogue-family keys (skips 'sys:...' system-level keys).
    """
    norm = _norm_cat(catalogue_num)
    if norm in lookup:
        return lookup[norm]
    prefix = re.sub(r"\d{3}$", "", norm)
    if prefix != norm and prefix in lookup:
        return lookup[prefix]
    for key, mat in lookup.items():
        if key.startswith("sys:"):
            continue
        if norm.startswith(key) or key.startswith(norm):
            return mat
    return None


# ===========================================================================
# CLI
# ===========================================================================


def main():
    parser = argparse.ArgumentParser(
        description="Knee Implant Reference Library Pipeline v3"
    )
    parser.add_argument(
        "--gudid",
        default=None,
        help="Path (or glob) to existing GUDID CSV — skips GUDID download",
    )
    parser.add_argument(
        "--mdall",
        default=None,
        help="Path to existing MDALL CSV — skips MDALL download",
    )
    parser.add_argument(
        "--merged",
        default=None,
        help="Path to existing merged GUDID+MDALL CSV — skips download + merge",
    )
    parser.add_argument(
        "--cjrr",
        default=None,
        help="Path to existing CJRR-filtered CSV — skips all earlier stages",
    )
    parser.add_argument(
        "--csv",
        default=None,
        help="Path to NZ medical devices supplemental CSV",
    )
    parser.add_argument(
        "--eifu-dir",
        default="implant_eIFUs",
        help="Directory containing eIFU PDFs (default: implant_eIFUs)",
    )
    parser.add_argument(
        "--no-cjrr-filter",
        action="store_true",
        help="Skip CJRR manufacturer filter",
    )
    parser.add_argument(
        "--no-gmdn-filter",
        action="store_true",
        help="Skip GMDN instrument filter after GUDID download",
    )
    parser.add_argument(
        "--download-eifus",
        action="store_true",
        help="Automatically download missing eIFU PDFs before enrichment "
        "(uses eifu_downloader.py; requires Selenium)",
    )
    parser.add_argument(
        "--eifu-headless",
        action="store_true",
        help="Run eIFU browser downloads in headless mode (no visible window)",
    )
    parser.add_argument(
        "--no-eifus",
        action="store_true",
        help="Skip all eIFU activity — no downloading, no PDF extraction "
        "(useful for quick runs or when Selenium is unavailable)",
    )
    parser.add_argument(
        "--no-gudid-api",
        action="store_true",
        help="Skip GUDID deviceDescription API enrichment "
        "(useful for offline runs; cached results still used if available)",
    )
    parser.add_argument(
        "--gudid-desc-cache",
        default="gudid_downloads/device_descriptions_cache.json",
        help="Path to the GUDID description JSON cache file",
    )
    parser.add_argument(
        "--output-dir",
        default="outputs",
        help="Output directory (default: outputs)",
    )
    args = parser.parse_args()

    pipeline = KneeImplantPipeline(
        medical_devices_csv_path=args.csv,
        eifu_directory=args.eifu_dir,
        apply_cjrr_filter=not args.no_cjrr_filter,
        apply_gmdn_filter=not args.no_gmdn_filter,
        download_eifus=args.download_eifus,
        download_eifus_headless=args.eifu_headless,
        skip_eifus=args.no_eifus,
        enrich_gudid_descriptions=not args.no_gudid_api,
        gudid_description_cache=args.gudid_desc_cache,
        gudid_csv=args.gudid,
        mdall_csv=args.mdall,
        merged_csv=args.merged,
        cjrr_csv=args.cjrr,
        output_dir=args.output_dir,
    )

    df = pipeline.run()
    print(
        f"\n[OK] Pipeline complete — {len(df):,} records written to {args.output_dir}/"
    )
    return df


if __name__ == "__main__":
    main()
